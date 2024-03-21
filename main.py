# https://colab.research.google.com/drive/1afriNsVrlMUDGTGwCAGgWumxZ0FTHuLe#scrollTo=H9Ws6ySA7em9
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import pandas as pd
import numpy as np
import sys, os, datetime, math
import openpyxl, subprocess
import json, requests, socket, re
from bs4 import BeautifulSoup
from urllib.request import urlopen
from urllib.error import HTTPError, URLError
import matplotlib
import matplotlib.pyplot as plt
from  matplotlib import font_manager as fm

class DiagGantt(QDialog):
    def __init__(self, view):
        super().__init__()    
        self.setWindowTitle('Выбор варианта')
        self.resize(300,180)
        layout = QVBoxLayout(self)

        select_label = QLabel('Вариант диаграммы')
        layout.addWidget(select_label)
        variants = ['SeriaShort', 'Seria', 'Cause', 'Equip']
        self.select_var = QComboBox()
        self.select_var.addItems(variants)
        layout.addWidget(self.select_var)
        btn_ok = QPushButton('ОК')
        btn_ok.clicked.connect(self.ok_clicked)
        layout.addWidget(btn_ok)
        btn_cancel = QPushButton('Cancel')
        btn_cancel.clicked.connect(self.cancel_clicked)
        layout.addWidget(btn_cancel)
    
    def ok_clicked(self):
        self.accept()

    def cancel_clicked(self):
        self.reject()

class DiagSignup(QDialog):
    def __init__(self, view):
        super().__init__()
        self.setWindowTitle('Авторизация')
        self.resize(300,180)
        layout = QVBoxLayout(self)
        login = QLabel('Логин')
        layout.addWidget(login)
        self.fio_combo = QComboBox()
        self.fio_combo.addItems(view.envir['users'])
        layout.addWidget(self.fio_combo)
        passw_label = QLabel('Пароль')
        layout.addWidget(passw_label)
        self.passw = QLineEdit()
        self.passw.setEchoMode(QLineEdit.Password) # Установка режима отображения символов заменителей
        layout.addWidget(self.passw)
        btn_ok = QPushButton('ОК')
        btn_ok.clicked.connect(self.ok_clicked)
        layout.addWidget(btn_ok)
        btn_cancel = QPushButton('Cancel')
        btn_cancel.clicked.connect(self.cancel_clicked)
        layout.addWidget(btn_cancel)
    
    def ok_clicked(self):
        self.accept()

    def cancel_clicked(self):
        self.reject()

class View(QMainWindow):
    btn_pressed = pyqtSignal(str) # Сигнал для передачи текста кнопки

    def __init__(self):
        super().__init__()
        self.envir = {'kmo_on':True,
        'noplaning_direct_on':False,
        'accuracy_plan':0.8,
        'fond_day_center':7.5,
        'adjust_equip':1,
        'works_count':19,
        'my_graf':'./График.xlsx',
        'my_setting':'./setting.cfg',
        'url_1c_login':'https://www.flim.ru/api/quality/login',
        'url_1c':'https://www.flim.ru/api/quality?format=json',
        'url_calendar':'https://www.consultant.ru/law/ref/calendar/proizvodstvennye/',
        'users':['Козин А.А.', 'Калайджи Ф.Н.', 'Волянская Е.Н.'],
        'kolvo_smen':{'3362 ТРЦ с ЧПУ LT-52':2,
                '4433 ТАПТ с ЧПУ MAZAK QT200MA':2,
                '3219 ФВОЦ Vcenter-70':1,
                '2032 ТРЦ с ЧПУ SL-20THE':1,
                'Верстак слесарный ЦМО №1':3,
                '4633 Станок нарезки кода ключа CRYPTEX (один шпиндель)': 1.4,
                    },
        'sort_order':{'sort':['YearSeria', 'ExecutionTime', 'Direct', 'SeriaShort', 'Operation'], 
                      'order':[True, True, True, False, True]},
                }
        if os.path.exists(self.envir['my_setting']):
            try:
                config = self.load_config()
                if config  is not None:
                    self.envir = config
            except:
                print('Файл с конфигурацией записан с ошибками')
        self.setWindowTitle("Планирование производства")
        self.pefix_path = 'Текущий путь: '
        self.pefix_file = 'Текущий файл: '
        self.folder_path = self.get_current_directory()
        self.file_name = 'Файл не выбран'
        # self.key_pattern = "3323 232 323 2212"
        self.login_str = None
        self.psw_str = None
        self.path_full = None
        self.session_1c = None
        self.open_list_file()
        self.initUI()
        self.create_tree()

    def initUI(self):
        tab_widget = QTabWidget()
        self.setCentralWidget(tab_widget)

        file_tab = QWidget()
        root_layout = QVBoxLayout(file_tab)
        tree_layout = QHBoxLayout()
        btn_layout = QVBoxLayout()

        self.path_label = QLabel(self.pefix_path + self.folder_path)
        root_layout.addWidget(self.path_label)

        self.file_label = QLabel(self.pefix_file + self.file_name)
        root_layout.addWidget(self.file_label)

        root_layout.addWidget(QLabel())

        root_layout.addLayout(tree_layout)
        tree_layout.addLayout(btn_layout)

        self.treeview = QTreeWidget()
        tree_layout.addWidget(self.treeview)
        self.treeview.itemClicked.connect(self.tree_callback)

        self.btn_path = QPushButton("Открыть в проводнике")
        btn_layout.addWidget(self.btn_path)
        self.btn_path.pressed.connect(lambda: self.btn_pressed.emit(self.btn_path.text()))#self.draft_path)

        self.btn_open_xls = QPushButton("Открыть файл")
        btn_layout.addWidget(self.btn_open_xls)
        self.btn_open_xls.pressed.connect(lambda: self.btn_pressed.emit(self.btn_open_xls.text()))

        self.btn_create_pattern = QPushButton("Загрузить данные 1С")
        btn_layout.addWidget(self.btn_create_pattern)
        self.btn_create_pattern.pressed.connect(lambda: self.btn_pressed.emit(self.btn_create_pattern.text()))
        # self.btn_create_pattern.pressed.connect(self.create_xls)

        self.btn_create_ms = QPushButton("Создать график")
        btn_layout.addWidget(self.btn_create_ms)
        self.btn_create_ms.pressed.connect(lambda: self.btn_pressed.emit(self.btn_create_ms.text()))      

        self.btn_create_g = QPushButton("Диаграмма Ганта")
        btn_layout.addWidget(self.btn_create_g)
        self.btn_create_g.pressed.connect(lambda: self.btn_pressed.emit(self.btn_create_g.text()))      

        self.btn_exit = QPushButton("Выход")
        btn_layout.addWidget(self.btn_exit)
        self.btn_exit.clicked.connect(self.close)

        setting_tab = QWidget()
        setting_layout = QVBoxLayout(setting_tab)
        setting_layout.addWidget(QLabel("Параметры планирования"))

        self.setting_plan = QTextEdit()
        setting_layout.addWidget(self.setting_plan)
        formatted_json = json.dumps(self.envir, indent=4, ensure_ascii=False).encode('utf-8').decode('utf-8')
        self.setting_plan.setPlainText(formatted_json)
        self.setting_plan.textChanged.connect(lambda: self.btn_pressed.emit('Изменть конфиг'))

        self.save_config = QPushButton("Нет изменений конфигурации")
        setting_layout.addWidget(self.save_config)
        self.save_config.pressed.connect(lambda: self.btn_pressed.emit('Сохранить конфиг')) 
        
        help_tab = QWidget()
        help_layout = QVBoxLayout(help_tab)
        help_msg = QTextEdit()
        help_layout.addWidget(help_msg)
        help_msg.setPlainText(self.text_help())

        tab_widget.addTab(file_tab, "Выбор файла")
        tab_widget.addTab(setting_tab, "Настройки")
        tab_widget.addTab(help_tab, "Справка")
        self.check_status()

    def get_current_directory(self):
        file_path = './main.py'
        path = os.path.abspath(file_path)
        return os.path.dirname(path)
    
    def load_config(self):
        my_setting = self.envir['my_setting']
        with open(my_setting, 'r') as file:
            params = json.load(file)
        return params
  
    def text_help(self):
        text = """
Версия 0.0.0.1
Программа предназначена для создания графика загрузки производственного оборудования.

Текущий путь - метка отображающая директорию, в которой будет создан файл График.xlsx. По умолчанию выбирана директория, содержащая исполняемый код. 

Текущий файл - метка отображающая полное имя выбранного файла в списке Файл.

Открыть в проводнике - команда открыть в проводнике Windows содержимого директрии Текущий путь.

Открыть файл - команда открыть для редактирования Текущий файл.

Загрузить данные 1С - команда загрузить данных из 1С. Требуется указать учетную запись и пароль доступа для сотрудников на сайт wwww.flim.ru

Создать график - команда создать таблицу загрузки оборудования на основе данных 1С с учетом Параметров планирования закладки Настройки.

Диаграмма Ганта - команда создать визуализацию графика загрузки оборудования. Для создания требуется выбрать поле группировки:
    SeriaShotr - номер наряда,
    Seria - номер серии номенклатуры,
    Cause - номер документа основания наряда,
    Equip - наименование оборудования

Выход - команда завершения работы программы.

Назначение параметров планирования. 
Внимание! Значения параметров заполнять по образцу. Если после внесения изменений работа программы завершается с ошибкой, то для восстановления значений по умолчанию в рабочем каталоге удалите файл setting.cfg

После вненсения изменений в параметры планирования перед началом создания графика обязательно выполнить команду Сохранить изменения конфигурации

'kmo_on':true, использовать при расчете знчения коэффициента многостаночного обслуживания. Возможные значения true, false

'noplaning_direct_on': false, использовать при рсчете знчения приоритета Не планировать. Возможные значения true, false

'accuracy_plan':0.8, минимальный период планирования в часах. Для отделения десятичных использовать точку '.'

'fond_day_center':7.5, фонд рабочего вреени рабочего центра в часах. Для отделения десятичных использовать точку '.'

'adjust_equip':1, время настройки оборудования между партиями в часах. Для отделения десятичных использовать точку '.'

'works_count':13, количество работников производства

'url_1c':'https://www.flim.ru/api/quality?format=json', адрес RES API УПП 1С

'url_calendar':'https://www.consultant.ru/law/ref/calendar/proizvodstvennye/' адрес вебстраницы производственного календаря

'users':['Козин А.А.', 'Калайджи Ф.Н.', 'Волянская Е.Н.'], список для выбора пользователя УПП 1С

'kolvo_smen':{'3362 ТРЦ с ЧПУ LT-52':2,
                '4433 ТАПТ с ЧПУ MAZAK QT200MA':2,
                '3219 ФВОЦ Vcenter-70':1,
                '2032 ТРЦ с ЧПУ SL-20THE':1,
                'Верстак слесарный ЦМО №1':3,
                '4633 Станок нарезки кода ключа CRYPTEX (один шпиндель)': 1.4
                    } список оборудования с указанием количества смен. Для отделения десятичных использовать точку '.' 1.4 соответсвует работе оборудования по 10 часов
'sort_order': {
        'sort': [
            'YearSeria',
            'ExecutionTime',
            ''Direct',
            'SeriaShort',
            'Operation'
        ],
        'order': [
            true,
            true,
            true,
            false,
            true
        ]
    } сортровка данных 1С, sort очередь сотрировки по полям, order порядок сортиовки true по возрастанию, false по убыванию
"""
        return text

    def create_xls(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'CrossTable'
        file_name = 'Шаблон МС.xlsx'
        file_path = os.path.join(self.folder_path, file_name)
        columns = ['K1', 'K2', 'K3', 'K4', 'K5']
        rows = {1: ['x', '', '', '', '', ], 2: ['x', '', '', '', '', ], 3: ['x', '', '', '', '', ], 4: ['x', '', '', '', '', ], 5: ['x', '', '', '', '', ]}

        # Запись заголовков столбцов
        for col_num, col_value in enumerate(columns, start=2):
            ws.cell(row=1, column=col_num, value=col_value)

        # Запись данных строк
        for row_num, row_values in rows.items():
            ws.cell(row=row_num+1, column=1, value=row_num)
            for col_num, col_value in enumerate(row_values, start=2):
                ws.cell(row=row_num+1, column=col_num, value=col_value)

        # Сохранение файла
        wb.save(file_path)
        self.open_list_file()
        self.create_tree()
    
    def check_status(self, msg_txt = "Готов к работе"):
        self.statusBar().showMessage(msg_txt)

    def tree_callback(self, item:QTreeWidgetItem, column):
        selected_item = item
        parent = item.parent()
        if parent is None:
            self.file_name = selected_item.text(0)
        else:
            self.file_name = parent.text(0)

        self.file_label.setText(self.pefix_file + self.file_name)
        self.path_full = os.path.join(self.folder_path, self.file_name)
        # is_valid_file = self.check_sheet_exists(self.path_full)
        # if is_valid_file:
        #     self.check_status()
        # else:
        #     self.check_status(f'В книге нет листа CrossTable')
        #     self.path_full = None

    def check_sheet_exists(self, file_path):
        if 'Файлов не обнаружено' in file_path:
            return False
        else:
            sheet_name = "CrossTable"
            wb = openpyxl.load_workbook(file_path)
            sheet_names = wb.sheetnames
            if sheet_name in sheet_names:
                return True
            else:
                return False

    def open_list_file(self):
        # Получение списка файлов в папке
        self.file_list = {file:[] for file in os.listdir(self.folder_path) if file.endswith('.xlsx') and not file.startswith('~$')}
        if len(self.file_list) == 0:
            self.file_list = {'Файлов не обнаружено': ['Листов не обнаружено']}
        # else:
        #     for file_name in self.file_list:            
        #         # Полный путь к выбранному файлу
        #         file_path = os.path.join(self.folder_path, file_name)
        #         # Загрузка файла Excel
        #         wb = openpyxl.load_workbook(file_path)
        #         # Получение списка листов в wb
        #         self.file_list[file_name] = wb.sheetnames
    
    def create_tree(self):
        self.treeview.setHeaderLabels(['Файл'])
        self.treeview.clear()
        for file, sheets in self.file_list.items():
            file_item = QTreeWidgetItem(self.treeview, [file])
            for sheet in sheets:
                sheet_item = QTreeWidgetItem(file_item, [sheet])
                file_item.addChild(sheet_item)
            self.treeview.addTopLevelItem(file_item)

    def is_file_cheked(self):
        path = os.path.join(self.folder_path, self.file_name)
        if os.path.exists(path) and self.file_name != 'Файл не выбран':
            self.check_status(msg_txt="Готов к работе")
            return True
        else:
            self.check_status(msg_txt="Файл не выбран")
            return False

class Presenter:
    def __init__(self, model, view, gantt):
        self.data = None
        self.model = model
        self.view = view
        self.gantt = gantt
        view.btn_pressed.connect(self.press_btn) # Подключение сигнала к методу
        self.is_changed = True

    def press_btn(self, btn_text): # Исправлено здесь
        match btn_text:
            case "Создать график":
                print("Нажата кнопка:", btn_text)
                self.model.create_graf(self.view)
                self.gantt.df_graf = None
            case "Загрузить данные 1С":
                print("Нажата кнопка:", btn_text)
                is_ok = self.model.load_data_1c(self.view)
                if is_ok:
                    self.model.prepere_data()
                    print('Загрузка данных завершена успешно')
            case "Открыть файл":
                print("Нажата кнопка:", btn_text)
                if self.view.is_file_cheked() and self.view.path_full is not None:
                    subprocess.Popen(['start', '', self.view.path_full], shell=True)
            case "Открыть в проводнике":
                print("Нажата кнопка:", btn_text)
                self.view.check_status(msg_txt ='Обновляем информацию о пути')
                subprocess.Popen(f'explorer {self.view.folder_path}')
                self.view.check_status()
            case "Диаграмма Ганта":
                print("Нажата кнопка:", btn_text)
                win_diag = DiagGantt(self.view)
                match win_diag.exec_():
                    case QDialog.Accepted:
                        print('Нажата кнопка Ок')
                        select = win_diag.select_var.currentText()
                        print(select)
                        self.gantt.plt_gantt(select)
                    case QDialog.Rejected:
                        print('Нажата кнопка Cancel')
            case 'Изменть конфиг':
                if self.is_changed:
                    print('Изменена конфигурация')
                    self.view.save_config.setText('Сохранить изменения конфигурации')
                self.is_changed = True
            case 'Сохранить конфиг':
                print('Нажата кнопка Сохранить изменения конфигурации')
                self.is_changed = False
                self.view.save_config.setText('Нет изменений конфигурации')
                txt = self.view.setting_plan.toPlainText()
                dic = self.model.text_to_dict(txt)
                self.model.save_config(dic)
                self.view.envir = self.view.load_config()
                formatted_json = json.dumps(self.view.envir, indent=4, ensure_ascii=False)
                self.view.setting_plan.setPlainText(formatted_json)
                
class Model():
    def __init__(self, view):
        super().__init__()
        self.view = view
        self.df = None
        self.path = view.envir['my_graf']
        self.envir = view.envir

    def text_to_dict(self, text):
        try:
            data = json.loads(text)
            if isinstance(data, dict):
                return data
            else:
                print("Ошибка: Текст не может быть преобразован в словарь.")
        except json.JSONDecodeError as e:
            print(f"Ошибка: Неверный формат JSON. {str(e)}")
    
    def save_config(self, params):
        my_setting = self.view.envir['my_setting']
        with open(my_setting, 'w') as file:
            json.dump(params, file)

    def open_diag_signup(self, view):
        print('Требуется авторизация...')
        view.check_status('Требуется авторизация...')
        win_diag = DiagSignup(view)
        match win_diag.exec_():
            case QDialog.Accepted:
                print('Нажата кнопка Ок')
                view.login_str = win_diag.fio_combo.currentText()
                view.psw_str = win_diag.passw.text()
                is_singn_1c = self.signup_1c(view)
                if is_singn_1c:
                    print('Пароль верный')
                    return True
                else:
                    print('Повторите ввод пароля')
                    view.check_status()
                    return False
                
            case QDialog.Rejected:
                print('Нажата кнопка Cancel')
                view.check_status()
                return False

    def signup_1c(self, view):
        print('Идет проверка прав пользователя...')
        view.check_status('Идет проверка прав пользователя...')
        try:
            ses = requests.Session()
            res = ses.post(view.envir['url_1c_login'], 
                           data={'login':view.login_str, 
                                 'password':view.psw_str, 
                                 'format':'json'})
            j_obj = res.json()
            #флаг авторизации
            aut = j_obj['authorization']['login']
            if aut == 'true':
                view.session_1c = ses
                self.load_data_1c(view)
                return True
            else:
                ses.close()
                print('Неверный пароль')
                self.open_diag_signup(view)
                return False
        except:
            print(TypeError, ValueError)
            aut = 'false'
            ses.close()
            print('Нет ответа от сервера. Попробуйте позже')
            self.open_diag_signup(view)
            return False
    
    def load_data_1c(self, view):
        print('Идет загрузка данных...')
        view.check_status('Идет загрузка данных...')
        try:
            ses = view.session_1c
            res = ses.get('https://www.flim.ru/api/quality?format=json')
            res_json = res.json()
            df_first =  pd.DataFrame(res_json.values())
            df_next = pd.DataFrame(df_first.row[0])
            self.df = df_next
            print('Данные успешно загружены')
            view.check_status()
            # ses.close()   
        except:
            if view.session_1c is None:
                is_ok= self.open_diag_signup(view)
                if not is_ok:
                    return False
        return True
    
    def prepere_data(self):
        if not self.create_df():
            print('Ошибка парсинга')
        self.df_full.ExecutionTime =  self.df_full.ExecutionTime.apply(self.data_str2date)
        self.df_full.Time = self.df_full.Time.apply(self.replace_comma_dot_split)
        print('Сохраняем исходные данные в файл ...')
        self.write_df2xls(sheet_name='Данные с ПЗВ', df=self.df_full)
        self.min2hour()
        return
    
    def create_graf(self, view):
        print('Создаем график загрузки оборудования...')
        grafic = Grafic(view)
        return True
    
    def print_counter(self):
        bar_length = 20
        percent = self.current_row/self.all_row
        bar = "[" + "≡" * int(percent * bar_length+1) + " " * (bar_length - int(percent * bar_length+1)) + "]"
        msg = f"\r{bar} {int(percent * 100)}%"
        sys.stdout.write('\r' + ' ' * len(msg) + '\r')
        sys.stdout.flush()
        print(msg, end="")
        self.current_row += 1
        return 
    
    def min2hour(self):
        self.df_full.reset_index(drop = True, inplace= True)
        # col = ['SeriaShort', 'Seria', 'Cause', 'Material', 'Direct', 'Equip', 'Operation',
        #         'ExecutionTime']#SeriaShort Number
        df = pd.DataFrame()
        self.df_full['Flag'] = np.nan
        self.all_row = len(self.df_full)
        self.current_row = 0
        print('Суммируем ПЗВ с основным временем ...')
        while len(self.df_full[self.df_full.Flag.isna()].head(1)) > 0:
            df_conc = self.df_full.loc[self.df_full.Flag.isna()].head(1)
            time_sum = self.df_full[(self.df_full.SeriaShort == df_conc.SeriaShort.values[0]) &
                                (self.df_full.Operation == df_conc.Operation.values[0]) &
                                (self.df_full.Equip == df_conc.Equip.values[0])].Time.sum()
            df_conc.loc[:,'Time'] = time_sum
            df = pd.concat([df, df_conc.loc[:].copy()],ignore_index = True)
            df_filter = self.df_full.loc[(self.df_full.SeriaShort == df_conc.SeriaShort.values[0]) &
                            (self.df_full.Operation == df_conc.Operation.values[0]) &
                            (self.df_full.Equip == df_conc.Equip.values[0])].copy()
            self.df_full.loc[df_filter.index,'Flag'] = 1
            self.print_counter()

        df['Date'] = np.NaN # Обнуляем график
        df['Tday'] = np.NZERO # Обнуляем норматив фонда рабочего времени
        df.Operation = df.Operation.astype('int')
        df.Tday = df.Tday.astype('int')
        df.Time = df.Time.astype('float')
        df['KMO']	= np.NaN # Обнуляем коэффициент совместного использования
        df['Parent'] = df.Cause.apply(self.extract_words) #
        # Удаляем ошибки по полю Direct из-за добавления данных  по кооперации
        print('\nКорректируем данные по кооперации...')
        seria_list = df.SeriaShort.unique()
        self.all_row = len(seria_list)
        self.current_row = 0
        for ser in seria_list:
            direct_uno = df.Direct.loc[(df.SeriaShort == ser) & (df.Equip != '3001 Кооперация')].iloc[0]
            df_filter = df.loc[(df.SeriaShort == ser)].copy()
            df.loc[df_filter.index,'Direct'] = direct_uno
            self.print_counter()
        print('\nКорректирка завершена')
        print('Сортируем таблицу по данным из трех колонк...')
        list_sort = self.envir['sort_order']['sort']
        list_order = self.envir['sort_order']['order']
        df.sort_values(list_sort, ascending=list_order, inplace=True)
        print('Удаляем дубликаты, оставляем текущее оборудование...')
        df.drop_duplicates(subset=['SeriaShort', 'Operation'], inplace=True, ignore_index = True)
        print('Переводим минуты в часы...')
        df.Time = round(df.Time/60, 2) # Переводим минуты в часы, точность страдает
        df.reset_index(drop= True , inplace= True )#Переписываем индекс
        print('Выгружаем результы в файл...')
        self.write_df2xls(sheet_name='Данные в часах', df=df)#Выгрузка результатов в файл

        # self.df_copy = df.copy()#Создаем копию таблицы, чтобы не тратить время на загрузку таблицы из файла

        return True
    
    def extract_words(self, text):
        #
        # Программа выделяет из текста номер документа
        # Используем регулярное выражение для поиска слов
        #
        re_word = re.search(r'[А-ЯЁ\d]{2}\d{9}', text)
        result = ''

        if re_word is not None and 'Наряд' in text:
            result = re_word.group()

        return result

    def replace_comma_dot_split(self, x):
        #
        #Преобразование строки в число с плавающей точкой
        #
        return float('.'.join(str(x).replace('\xa0', '').split(',')))

    def data_str2date(self, data_text):
        """
        Конвертируем текст в дату
        """

        day_int = int(data_text[0:2])
        month_int = int(data_text[3:5])
        year_int = int(data_text[6:10])
        data_data = datetime.date(year_int,month_int,day_int)
        return data_data

    def create_df(self):
        if self.df is None:
            print('Загрузите данные')
        try:
            df_next = self.df
            self.df_full = pd.DataFrame()
            print('Идет парсинг данных...')
            for i in range(len(df_next.Value)):
                self.df_full = pd.concat([pd.DataFrame(df_next.Value[i]).T , self.df_full.loc[:].copy()])
            # df_ful = pd.concat([pd.DataFrame(df_next.Value[j]).T for j in range(len(df_next.Value))], ignore_index=True).append(df_ful, ignore_index=True)
            # 'YearSeria' год создания наряда
            # 'SeriaShort' номер наряда
            # 'Seria' серия номенклатуры
            # 'Cause' документ основание для наряда
            # 'Material' продукция
            # 'Direct' приоритет
            # 'Equip' оборудование или группа заменяемости
            # 'Operation' номер операции технологичнского процесса
            # 'Time' трудоемкость технологической операции
            # 'ExecutionTime' плановая дата выполнения наряда
            # 'Date' номер рабочего дня
            # 'Tday' плановый дневной фонд времени
            # 'KMO' коэффициент многастоночного обслуживания
            # 'Parent' номер наряда родителя
            print('Парсинг данных завершен')
            self.df_full.columns = ['YearSeria', 'SeriaShort', 'Seria', 'Cause',
                                'Material', 'Direct', 'Equip',
                                'Operation', 'Time', 'ExecutionTime']# , 'DocBase'
            return True
        except:
            return False

    def write_df2xls(self, sheet_name, df):
        with pd.ExcelWriter(self.path, mode="a", engine="openpyxl", if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name) 
        
        return True

class Grafic:
    def __init__(self, view):
        super().__init__()
        self.view = view
        print('Готовим данные для создания графика...')
        self.df = pd.read_excel(self.view.envir['my_graf'], sheet_name='Данные в часах')
        self.envir = view.envir
        self.pay_roll = 7.5*self.envir['works_count']
        self.create_dic_gzrc()
        self.create_df_kmo()
        print('Создаем график...')
        self.create_graf()
        print('График готов')

    def create_dic_gzrc(self):
        self.df_gzrc = pd.read_excel('./GZRC.xlsx')
        self.df_gzrc.columns = ['Group',  'WC1', 'WC2',
            'WC3', 'WC4', 'WC5',
            'Kint1',
            'Kint2',
            'Kint3',
            'Kint4', 'Kint5']
        cnt = len(self.df_gzrc)
        self.dic_gzrc = dict()
        for i in range(cnt):
            self.dic_gzrc[self.df_gzrc.iloc[i,0]] = [str(self.df_gzrc.iloc[i,x]) for x in range(1,5) if str(self.df_gzrc.iloc[i,x]) != 'nan' ]

    def create_df_kmo(self):
        # Создаем таблицу КМО
        self.df_kmo = pd.DataFrame()
        for i in range(1,6):
            col1 = 'WC' + str(i)
            col2 = 'Kint' + str(i)
            df_kmo_temp = self.df_gzrc.loc[ : , [col1,col2] ].copy()
            df_kmo_temp.columns =['WC','KMO']
            self.df_kmo = pd.concat([self.df_kmo, df_kmo_temp.loc[:].copy()])
        self.df_kmo = self.df_kmo[~(self.df_kmo.WC.isna() | self.df_kmo.KMO.isna())]
        self.df_kmo.drop_duplicates(keep='first', inplace=True)
        self.df_kmo.sort_values('WC', inplace=True)
        self.df_kmo.reset_index(inplace=True, drop = True)

    def print_counter(self):
        bar_length = 20
        percent = self.current_row/self.all_row
        bar = "[" + "≡" * int(percent * bar_length+1) + " " * (bar_length - int(percent * bar_length+1)) + "]"
        msg = f"\r{bar} {int(percent * 100)}%"
        sys.stdout.write('\r' + ' ' * len(msg) + '\r')
        sys.stdout.flush()
        print(msg, end="")
        self.current_row += 1
        return 

    def choose_child_idx(self, idx_in):
        #
        # Программа просматривает всех потомков и выбирает индекс строки первого потомка
        # в последнем поколении
        #
        parent = self.df.loc[idx_in, 'SeriaShort'].values[0]
        if self.noplan_on: 
            idx_out = list(self.df[(self.df.Parent == parent) & (self.df.Date.isna())].head(1).index)
        else:
            idx_out = list(self.df[(self.df.Parent == parent) & (self.df.Date.isna()) & (self.df.Direct.notna())].head(1).index)

        if len(idx_out) > 0:
            self.choose_child_idx(idx_out)
        else:
            idx = idx_in
        return idx 

    def wait_time(self, item_gt = 'Seria', seria = '07.232.4Z'):
        #
        #Создаем из списка в вида [0,1,6] кортеж ([6, 4])
        #
        wait_list = []# Список ожидания, это результат программы
        # Список ожидания:До какого числа ожидает, сколько дней ожидает, оборудование, оборудование
        temp_list =[]# Временный список

        if len(self.df_graf[self.df_graf[item_gt] == seria]) ==0:
            return wait_list

        # Выбираем даты (индекс) и загрузку (колонка)
        table = self.df_graf[['Date', 'Time']][self.df_graf[item_gt] == seria].groupby(by="Date").sum()
        days = table.index# Получаем список дат
        for i in range(len(days)):# Перебираем список дней
            if i == 0:# проверяем первый элемент списка
                if days[i] == 1:# Если = 1
                    pass# то ничего не делаем
                else:# Иначе
                    # Получаем нименование оборудования или серии
                    equip_item = self.df_graf[(self.df_graf[item_gt] == seria) & (self.df_graf['Date'] == days[i])].Equip.unique()[0]
                    # Зполняем список ожидания
                    wait_list.append((days[i], days[i] - 1, seria, equip_item))#
            elif days[i] - days[i-1] > 1:# Если интервал больше единицы
                equip_item = self.df_graf[(self.df_graf[item_gt] == seria) & (self.df_graf['Date'] == days[i])].Equip.unique()[0]
                wait_list.append((days[i], days[i] - days[i-1]-1, seria, equip_item))
            elif i == len(days)-1:# если конец списка
                pass
            else:
                pass
        return wait_list
    
    def creat_go_list(self, item_gt = 'Equip', seria = 'Верстак слесарный ЦМО №1', gzrc_index = 0):
        
        free_day = [] # Начало свободного интервал
        free_end = [] # Конец свободго интервала
        free_time = [] # Свободное время в часах
        free_equip = [] # Свободное оборудование

        equip_list = self.wait_time(item_gt, seria)
        # Список ожидания:До какого числа ожидает, сколько дней ожидает, оборудование, оборудование
        for i in range(len(equip_list)):
            free_end.append(equip_list[i][0])
            free_day.append(equip_list[i][0]-equip_list[i][1])
            free_time.append(equip_list[i][1]*self.envir['fond_day_center'])# ((equip_list[i][0]-equip_list[i][1])*fond_day_center)
            free_equip.append(equip_list[i][2])

        # and merge them by using zip().
        list_tuples = list(zip(free_day, free_time, free_equip, free_end))

        # Converting lists of tuples into
        # pandas Dataframe.
        df_go_list = pd.DataFrame(list_tuples, columns=['FreeDay', 'FreeTime', 'Equip', 'FreeEnd'])
        df_go_list['GzrcIndex'] = gzrc_index
        return df_go_list
       
    def draft_equipt(self, s, last_seria, free_day):
        #
        # Выбираем оборудование на основании данных о его загрузке
        # Используем построение графика по принципу Ранний старт
        # dic_gzrc_temp = dic_gzrc.copy()
        # s = 0
        # last_seria = ''
        # free_day = 0
        # df_graf = pd.concat([df_graf, df[rw:rw+1]], ignore_index = True)
        #
        # index_col_time = self.df.columns.get_loc('Time')#Создаем константы для имени колонки таблицы
        # index_col_date = self.df.columns.get_loc('Date')#Создаем константы для имени колонки таблицы
        index_col_equip = self.df.columns.get_loc('Equip')#Создаем константы для имени колонки таблицы
        index_col_seria = self.df.columns.get_loc('SeriaShort')#Создаем константы для имени колонки таблицы
        index_col_material = self.df.columns.get_loc('Material')#Создаем константы для имени колонки таблицы
        index_col_op = self.df.columns.get_loc('Operation')#Создаем константы для имени колонки таблицы
        # index_col_kmo = self.df.columns.get_loc('KMO')#Создаем константы для имени колонки таблицы
        # index_col_parent = self.df.columns.get_loc('Parent')#Создаем константы для имени колонки таблицы
        gzrc_index = 0
        df_col = ['FreeDay', 'FreeTime', 'Equip', 'FreeEnd', 'GzrcIndex']
        seria = self.df_graf.iloc[s, index_col_seria] # Серия
        op = self.df_graf.iloc[s, index_col_op] # Операция
        prod = self.df_graf.iloc[s, index_col_material]# Продукция
        equip_op = self.df_graf.iloc[s, index_col_equip] # Оборудование
        current_seria = str(prod) + ' ' + str(seria) + ' ' + str(op)

        if last_seria != current_seria:
            last_seria = current_seria
            equip_gzrc = self.df_graf.iloc[s, index_col_equip] # Оборудование, мб ГЗРЦ
            if equip_gzrc == '3001 Кооперация':
                gzrc_index = 0
                free_day = self.df_graf.Date.loc[(self.df_graf.SeriaShort == seria)].max()
                # df_graf.loc[(df_graf.SeriaShort == seria)]
            else:
                # Создаем пустую таблицу для сбора информации о свободных днях оборудования
                df_go_list = pd.DataFrame() # columns=df_col
                # Получаем информацию из выбранной строки
                equip_op = ''  # Стираем информацию об операции
                # Считаем трудоемкость операции в текущей серии ['Time']
                op_trud = self.df_graf.Time.loc[(self.df_graf['SeriaShort'] == seria) & (self.df_graf['Operation'] == op)].sum()#Seria
                op_trud = round(op_trud, 2)

                #Вычисляем последний день серии
                last_day_temp = self.df_graf.Date.loc[self.df_graf['SeriaShort'] == seria].max()#Seria
                # last_day_temp = last_day_temp['Date']#Вычисляем последний день серии
                if math.isnan(last_day_temp): # == None:
                    last_day = 1 # Если дней нет, то последний день серии = 1
                else:
                    last_day = last_day_temp # Иначе последний день серии = максимльный день серии
                # Подбираем оборудование
                if equip_gzrc in self.dic_gzrc and equip_gzrc != '3001 Кооперация': # Если текущая ГЗРЦ существует в списке
                    dic_gzrc_uno = {equip_gzrc : self.dic_gzrc[equip_gzrc]}
                else:
                    dic_gzrc_uno = {equip_gzrc : [equip_gzrc]}

                gzrc_len = len(dic_gzrc_uno[equip_gzrc]) # Считаем длину списка группы

                for i in range(gzrc_len): # Перебираем список группы по порядку, создаем df_go_list
                    # i = 0
                    # Ищем дырки в загрузке оборудования
                    item_gt = 'Equip'
                    seria = dic_gzrc_uno[equip_gzrc][i]
                    gzrc_index = i

                    df_go_list_temp = self.creat_go_list(item_gt, seria, gzrc_index)
                    # Проверяем количество записей для выбранного оборудования в формируемом графике

                    count_equip = self.df_graf.Equip.loc[self.df_graf['Equip'] == dic_gzrc_uno[equip_gzrc][i]].count()


                    if count_equip == 0:  # Если в графике записей нет
                        equip_free_day = 0  # то устанавливаем первый свободный день оборудования = 0
                    else:
                        # иначе первый свободный день  = макс дню в формируемом графике для выбранного оборудования
                        equip_free_day = self.df_graf.Date.loc[self.df_graf['Equip'] == dic_gzrc_uno[equip_gzrc][i]].max()
                        if math.isnan(equip_free_day): # == None дни для оборудования не запланированы
                            equip_free_day = 0

                    if len(df_go_list_temp) == 0: # Если дырок в загрузке оборудования нет
                        # то устанавливаем последний свободный день = макс дню в формируемом графике
                        last_day_all = self.df_graf.Date.max() #
                        if math.isnan(last_day_all): # == None:
                            last_day_all = 0
                        # формируем таблицу о свободных днях оборудования по умолчанию в формате
                        # 'FreeDay', 'FreeTime', 'Equip', 'FreeEnd', 'GzrcIndex'
                        df_data = [[equip_free_day, op_trud , dic_gzrc_uno[equip_gzrc][i], last_day_all, i]]
                        df_go_list_temp = pd.DataFrame(data =df_data, columns=df_col)
                    # Записываем информацию о свободных днях по всему списку оборудования в одну таблицу
                    # if df_go_list.shape[0] == 0:
                    #     df_go_list = df_go_list_temp.copy()
                    # else:
                    df_go_list = pd.concat([df_go_list, df_go_list_temp.loc[:].copy()])


                # исключаем из списка о свободных днях информацию не удовлетворяющую условиям
                df_for_if = df_go_list.loc[df_go_list['FreeEnd'] >= last_day]
                df_for_if = df_go_list.loc[df_go_list['FreeDay'] < last_day]
                df_for_if = df_go_list.loc[df_go_list['FreeTime'] >= op_trud]

                # Если есть оборудование, удовлетворяющее условиям
                if len(df_go_list) != 0 and len(df_for_if) != 0:
                    #Сортируем таблицу по данным из трех колонк
                    # df_for_if.sort_values([ 'FreeDay', 'GzrcIndex'], inplace=True)
                    df_for_if = df_for_if.loc[df_for_if.sort_values(['FreeDay', 'GzrcIndex']).index]
                    df_for_if.reset_index(drop= True , inplace= True )#Переписываем индекс
                    equip_op = df_for_if.iloc[0] # Берем лучший результат оборудования из 0-й строки
                    gzrc_index = equip_op['GzrcIndex']
                    # Выбираем оборудование с меньшим первым свободным днем и большим приоритетом в ГЦРЗ
                    equip_op = equip_op['Equip']

                    free_day = df_for_if.iloc[0]
                    free_day = free_day['FreeDay'] # Выбираем первый свободный день
                else: # Если нет оборудования, удовлетворяющего условиям
                    equip_op = dic_gzrc_uno[equip_gzrc][0] # то выбираем оборудование первое в списке ГЦРЗ
                    gzrc_index = 0
                    free_day = -1  # устанавливаем первый свободный день

        return equip_op, free_day, current_seria, gzrc_index

    def mul_row(self, i, k, adj, dic_last, flag_gzrc):
        # 
        # Добавляем к таблице графика строки, содержащие машинное время не более точности плнарирования
        # Суммарное время машинного времени не должно быть выше трудоемкости выбранной операции
        # 
        index_col_time = self.df.columns.get_loc('Time')#Создаем константы для имени колонки таблицы
        df_row = pd.DataFrame() # Создаем пустую таблицу
        # делаем копию текущей строки таблицы
        df_temp = pd.DataFrame(self.df.loc[i].copy()).transpose()
        # df_temp = df.loc[i].copy()
        # current_time = df_temp.iloc[0, index_col_time]
        # Получаем текущее машинное время [0, index_col_time]
        current_time = df_temp.loc[i,'Time']
        #
        #
        # Добавить выбор времени наладки оборудования
        #
        # Если флаг = 0, значит оборудование уже налажено и работает
        if flag_gzrc:
            current_time += adj  # dic_last <<<<<<<<<<<<<<

        # Сверяем текущее машинное время с точностью планирования
        if current_time > self.envir['accuracy_plan']*k:
            # В 0 строке временной таблицы из одной строки устанавливаем машинное время равное точности планирования
            df_temp.iloc[0, index_col_time] = self.envir['accuracy_plan']*k
        # Сверяем суммарное машинное время временной из k строке рабочей таблицы с точностью планирования
        while current_time > self.envir['accuracy_plan']*k:
            # пока выполняется неравество уменьшаем текщее машинное время из k строке рабочей таблицы на величину точности планирования
            current_time -= self.envir['accuracy_plan']*k
            # объединяем строки временной таблицы
            df_row = pd.concat([df_row, df_temp.loc[:].copy()])

        df_temp.iloc[0, index_col_time] = current_time # записываем во временную таблицу из одной строки
        df_row = pd.concat([df_row, df_temp.loc[:].copy()]) # объединяем строки временной таблицы
        df_row['KMO'] = k
        if k != 1:
            df_row['Time'] = df_row['Time']/k
        df_row['Date'] = np.NaN
        # df_row['Df_Index'] = i

        return df_row, dic_last

    def check_fond_all(self, day, s, df_def):
        #
        # Контроль переполнения фонда времени указанного дня и указанного оборудования
        # Добавить контроль переполнения фонда времени по заказу в день
        # Добавить замену группы на конкретное оборудование из группы для всего заказа
        # Надо проверять было ли присвоение оборудования к данному заказу
        #  Выполнено.  Для каждого заказа проверять свобоное оборудование с 1го числа.
        #
        index_col_equip = self.df.columns.get_loc('Equip')#Создаем константы для имени колонки таблицы
        index_col_seria = self.df.columns.get_loc('SeriaShort')#Создаем константы для имени колонки таблицы
        index_col_kmo = self.df.columns.get_loc('KMO')#Создаем константы для имени колонки таблицы

        f_day = self.pay_roll
        f_equip = self.envir['fond_day_center']
        equip = df_def.iloc[s, index_col_equip] #Получаем в текущей строке наименование оборудования
        seria = df_def.iloc[s, index_col_seria] #Получаем в текущей строке наименование серии
        # Изменяем фонд времени в соответсвии со сменным графиком работы оборудования
        if equip in self.envir['kolvo_smen']:
            f_equip = self.envir['fond_day_center']*self.envir['kolvo_smen'][equip]

        if self.envir['kmo_on']:
            kmo = df_def.iloc[s, index_col_kmo] #Получаем в текущей строке кмо
        else:
            kmo = 1
        # ['Time'] Считаем фонд времени выбранного оборудования и текущего дня
        sum_equip = df_def.Time[(df_def['Equip'] == equip) & (df_def['Date'] == day)].sum()
        #['Time']Считаем фонд времени текущего дня
        sum_day = df_def.Time[df_def['Date'] == day].sum()
        #Seria#['Time'] Считаем фонд времени выбранной серии и текущего дня
        sum_seria = df_def.Time[(df_def['SeriaShort'] == seria) & (df_def['Date'] == day)].sum()

        if kmo != 1:
            if equip != '3001 Кооперация':
                check = sum_seria >= f_equip or sum_equip >= f_equip
            else:
                check = sum_seria >= f_equip
        else:
            if equip != '3001 Кооперация':
                check = sum_seria >= f_equip or sum_equip >= f_equip or sum_day >= f_day
            else:
                check = sum_seria >= f_equip

        return  check #Сравниваем фактический фонд времени с плановым
    
    def creat_graf_row(self, df_graf, s, free_day):
        #
        # Первая версия программы
        # Перебираем таблицу данных
        # По данным текущей строки ищем максимальный день для заказа
        # По данным текущей строки ищем максимальный день для оборудования
        # New По данным наряда, где SeriaShort==Parent ищем максимальный день
        # New Из трех полученных дат получаем максимальную
        # С этой даты начинаем заполнять даты для текущего заказа + оборудование
        # Контролируем трудоемкость текущего дня
        # Контролируем трудоемкость текущего оборудования
        # Если трудоемкость превышена увеличиваем день на 1
        #
        index_col_equip = self.df.columns.get_loc('Equip')#Создаем константы для имени колонки таблицы
        index_col_seria = self.df.columns.get_loc('SeriaShort')#Создаем константы для имени колонки таблицы
        index_col_kmo = self.df.columns.get_loc('KMO')#Создаем константы для имени колонки таблицы
        index_col_date = self.df.columns.get_loc('Date')#Создаем константы для имени колонки таблицы

        equip = df_graf.iloc[s, index_col_equip] #Получаем в текущей строке наименование оборудования
        seria = df_graf.iloc[s, index_col_seria] #Получаем в текущей строке наименование серии
        kmo = df_graf.iloc[s, index_col_kmo] #Получаем в текущей строке кмо
        # Код для parent НЕ удалять
        # parent = df_graf.iloc[s, index_col_parent] #Получаем в текущей строке parent

        if free_day == -1:
            # Получаем для выбранного оборудование максимальный календарный день
            max_day_equip = df_graf.Date[df_graf['Equip'] == equip].max()
            if math.isnan(max_day_equip) : # Проверяем, что дни для оборудования были назначены is None
                max_day_equip = 1 # Если не один день не назначен, то присваиваем переменной значение 1
            max_day_temp = [max_day_equip] # Записываем день в список
        else:
            max_day_temp = [free_day]
        #Получаем для выбранной серии максимальный календарный день
        max_day_seria = df_graf.Date[df_graf['SeriaShort'] == seria].max()
        if math.isnan(max_day_seria): # Проверяем что дни для оборудования были назначены
            max_day_seria = 1 #Если не один день не назначен, то присваиваем переменной значение 1
        max_day_temp.append(max_day_seria)#Добавляем день в список

        # Код для parent НЕ удалять
        # if math.isnan(parent):
        #   max_day_parent = 0 # Если не один день не назначен, то присваиваем переменной значение 0
        # max_day_parent = df_graf.Date[df_graf['Parent'] == parent].max() #Получаем для выбранного родителя максимальный календарный день
        # if math.isnan(max_day_parent): # Проверяем что дни для родителя были назначены
        #   max_day_parent = 0 # Если не один день не назначен, то присваиваем переменной значение 0
        # max_day_temp.append(max_day_parent)# Добавляем день в список

        max_day = max(max_day_temp) # Из трех дней выбираем максимальный
        df_graf.iloc[s, index_col_date] = max_day # Записываем номер дня в текущую строку таблицы
        # Увеличиваем плановый фонд с учетом оптимизации fond_day_center*. Замена 1 на mul_pay_roll_per_day[max_day-1], equips[equip]
        fond_mul_correction = self.envir['fond_day_center']*1*1
        if self.check_fond_all(max_day, s, df_graf): # Проверяем соблюдения дневного фонда и фонда оборудования
            while self.check_fond_all(max_day, s, df_graf): # Увеличиваем номер планируемого дня, пока не будут выполнены дневной фонд и фонд оборудования
                max_day += 1
                df_graf.iloc[s, index_col_date] = max_day

        return df_graf

    def calendar_up(self):
        #
        # Создаем календарь рабочих дней с сайта Консультант
        #
        def get_bs(url):
            """
            Программа проверяет
            доступность сервера,
            доступность страницы
            иначе выводит сообщение об ошибке
            """
            try:
                html_cont = urlopen(url, timeout=5.0)
                html_bs = html_cont.read()
                html_cont.close()
            except HTTPError as e:
                print('The page could not be found!')
                return -1
            except URLError as e:
                print(f'The server {url} could not be found!')
                return -1
            except socket.timeout:
                print(f'socket timed out - URL {url}')
                return -1
            return BeautifulSoup(html_bs, 'lxml')
            
        def int_to_0str(x):
            """
            Программа возвращает число в формате строки со значимым 0 в первой позиции
            """
            if x < 10:
                x = '0' + str(x)
            return str(x)

        def str_to_0str(x):
            lst = list(x)
            if lst[-1] == "*":
                x = lst[-5:-1]
                x = "".join(x)
            if len(x) == 1:
                x = '0' + x
            return x
        
        modul_dt = datetime.datetime
        dt_now = modul_dt.now()
        year_now = dt_now.year
        years = [year_now, year_now + 1]
        day_all = []
        for year_one in years:
            # year_one  = 2024
            str_url = 'https://www.consultant.ru/law/ref/calendar/proizvodstvennye/' + str(year_one)
            html_txt = get_bs(str_url)
            if html_txt == -1:
                break
                # file_path = '/content/drive/MyDrive/Flim/Kosultant'+str(year_one)+'.html'
                # html_bs = open(file_path,'r')
                # html_txt = BeautifulSoup(html_bs, 'lxml')
                # html_bs.close()
                # df_calendar = pd.DataFrame()
                # return df_calendar
            print(html_txt.find_all( 'h2')[0].text, end = ' ')
            table_list = html_txt.find_all('table', class_="cal")
            i = 1
            for table in table_list:
                # table = table_list[0]
                manth_name = table.find(class_="month").text
                day_list = table.find_all('td')
                for day in day_list:
                    # day = day_list[0]
                    if day['class'] == ['inactively']:
                        pass
                    elif day['class'] == ['holiday', 'weekend']:
                        date_lst = [str_to_0str(day.text), int_to_0str(i), str(year_one)[-2:]]
                        date_str = '/'.join(date_lst)
                        day_all.append([modul_dt.strptime(date_str, '%d/%m/%y'), date_str, 'праздничный'])
                    elif day['class'] == []:
                        date_lst = [str_to_0str(day.text), int_to_0str(i), str(year_one)[-2:]]
                        date_str = '/'.join(date_lst)
                        day_all.append([modul_dt.strptime(date_str, '%d/%m/%y'), date_str, 'рабочий'])
                    elif day['class'] == ['work']:
                        date_lst = [str_to_0str(day.text), int_to_0str(i), str(year_one)[-2:]]
                        date_str = '/'.join(date_lst)
                        day_all.append([modul_dt.strptime(date_str, '%d/%m/%y'), date_str, 'рабочий'])
                    elif day['class'] == ['preholiday']:
                        date_lst = [str_to_0str(day.text), int_to_0str(i), str(year_one)[-2:]]
                        date_str = '/'.join(date_lst)
                        day_all.append([modul_dt.strptime(date_str, '%d/%m/%y'), date_str,  'короткий'])
                    elif day['class'] == ['weekend']:
                        date_lst = [str_to_0str(day.text), int_to_0str(i), str(year_one)[-2:]]
                        date_str = '/'.join(date_lst)
                        day_all.append([modul_dt.strptime(date_str, '%d/%m/%y'), date_str,  'выходной'])
                    else:
                        print("Добавить elif для неучтенного класса", day['class'])
                i += 1
            print(len(day_all))

        df_calendar = pd.DataFrame(day_all)
        df_calendar.columns = ['DateTime', 'DateStr', 'Type']
        df_calendar.astype({'DateTime': 'datetime64[ns]'})
        date_now = modul_dt.today()
        date_now = date_now.combine(date_now.date(), date_now.min.time())
        df_calendar = df_calendar[df_calendar.DateTime >= date_now]
        df_calendar.reset_index(inplace= True, drop = True)
        return df_calendar
    
    def graf_join_calendar(self, x):
        #
        # Добавляем к графику даты из графика рабочих дней
        #
        if np.isnan(x):
            date_out = self.df_calendar.DateTime.max()
            print('Errore. Нет данных Date.')
        else:
            x = int(x)
            if x <= len(self.df_calendar):
                date_out = self.df_calendar.loc[x-1, 'DateTime']
            else:
                date_out = self.df_calendar.DateTime.max()
        return date_out
    
    def create_graf(self):
        
        fond_day_center = self.envir['fond_day_center']
        kmo_on = self.envir['kmo_on']
        self.noplan_on = self.envir['noplaning_direct_on']
        adjust_equip = self.envir['adjust_equip']
        # Инициилизируем переменные
        # self.df = self.df_copy.copy()#Создаем копию таблицы, чтобы не тратить время на загрузку таблицы из интернета
        self.df_graf = pd.DataFrame()
        # df_calendar = pd.DataFrame()

        date_x = 0
        # hold_parent_child = set()
        last_seria = ''
        free_day = 0
        gzrc_index = 0
        last_eq_mat = {}
        

        # index_col_time = self.df.columns.get_loc('Time')#Создаем константы для имени колонки таблицы
        # index_col_date = self.df.columns.get_loc('Date')#Создаем константы для имени колонки таблицы
        # index_col_equip = self.df.columns.get_loc('Equip')#Создаем константы для имени колонки таблицы
        index_col_seria = self.df.columns.get_loc('SeriaShort')#Создаем константы для имени колонки таблицы
        index_col_material = self.df.columns.get_loc('Material')#Создаем константы для имени колонки таблицы
        index_col_op = self.df.columns.get_loc('Operation')#Создаем константы для имени колонки таблицы
        # index_col_kmo = self.df.columns.get_loc('KMO')#Создаем константы для имени колонки таблицы
        # index_col_parent = self.df.columns.get_loc('Parent')#Создаем константы для имени колонки таблицы

        # отбор строк при соблюдении условий
        # данные строки не использовались для графика загрузки оборудования
        # наряд родитель включен в график загрузки оборудования
        # выбираем первую строку результата

        # Обрабатываем таблицу с исходными данными, пока не будет полностью заполнено
        # поле Date, которое содержит номер рабочего дня.

        if self.noplan_on:
            len_df = self.df.shape[0]
        else:
            len_df = self.df[self.df.Direct.notna()].shape[0]

        while True:
            if self.noplan_on:
                if self.df[self.df.Date.isna()].head(1).shape[0]==0:
                    break
            else:
                if self.df[self.df.Date.isna() & self.df.Direct.notna()].head(1).shape[0]==0:
                    break
            # Построчно обрабатываем исходные данные, в которых указан рабочий центр,
            # а не группа заменяемости рабочих центров,  и не заполнено поле Date
            df_no_gzrc = self.df[(~self.df.Equip.isin(self.df_gzrc.Group)) & (self.df.Date.isna())].head(1)

            if len(df_no_gzrc) > 0:
                # отбор строк при соблюдении условий
                # данные строк не использовались для графика загрузки оборудования
                # оборудование не является членом списка групп заменяемости
                # выбираем первую строку результата
                # Добавляем в график следующую строку исходных данных для рабочих центров,
                # взятых в работу

                self.df_graf = pd.concat([self.df_graf, df_no_gzrc.loc[:].copy()], ignore_index = True)
                current_index = list(df_no_gzrc.index)
                # сбрасываем флаг обработки списка группы заменяемости рабочих центров
                flag_gzrc = 0

            else:
                # Добавляем в график следующую строку исходных данных для групп рабочих
                # центров, ожидающих завершения предыдущих операций
                # Получаем индекс текущей строки
                if self.noplan_on:
                    curr_idx = list(self.df[self.df.Date.isna()].head(1).index)
                else:
                    curr_idx = list(self.df[self.df.Date.isna() & self.df.Direct.notna()].head(1).index)
                # получаем индекс строки первого потомка
                # в последнем поколении
                current_index = self.choose_child_idx(curr_idx)
                # Строку исходных данных с индексом current_index записываем во временную
                # таблицу
                df_yes_gzrc = self.df.loc[current_index].head(1)
                # Объединяем строки графика и временной талицы.
                # в python есть много команд для объединения таблиц, но pd.concat
                # рекомендована как лучшее решение, отсюда все заморочки

                if df_yes_gzrc.shape[0] > 0:
                    self.df_graf = pd.concat([self.df_graf, df_yes_gzrc.loc[:].copy()], ignore_index = True)
                # устанавливаем флаг = 1, что означает, что работаем со списком группы
                # заменяемости, а не с отдельным рабочим центром
                flag_gzrc = 1

            # Для строки с индексом current_index получаем значения
            # номер наряда
            seria = self.df.loc[current_index, 'SeriaShort'].values[0]
            # seria = df_graf[df_graf.Date.isna()].iloc[0, index_col_seria]# Серия df
            # Номер операции
            op = self.df.loc[current_index, 'Operation'].values[0]
            # op = df_graf[df_graf.Date.isna()].iloc[0, index_col_op]# Операция df
            # наименование продукции
            prod = self.df.loc[current_index, 'Material'].values[0]
            # prod = df_graf[df_graf.Date.isna()].iloc[0, index_col_material]# Продукция df
            # Вычиялем номер последней строки графика
            end = len(self.df_graf) - 1
            # запоминаем значение last_seria в переменной last_seria_mem
            last_seria_mem = last_seria

            # получаем данные об индексе оборудования в ГЗРЦ gzrc_index
            # draft_equipt возвращает бльше данных, но сейчас нам нужен только индекс
            # рабочего центра в списке оборудования в группе заменяемости рабочих центров
            equip_op, free_day, last_seria, gzrc_index = self.draft_equipt(end, last_seria, free_day) #df_graf,

            #
            self.df_graf = self.df_graf[0:end] # удаляем последнюю строку из графика
            bgn = len(self.df_graf)
            # если из списка ГЗРЦ для графика подходит рабочий центр с индексом отличным
            # от 0, то значит данное оборудование должно использоваться в многостаночном
            # режиме, и время загрузки оборудования определяется коэффициентом КМО
            kmo = 1.0
            if gzrc_index != 0:
                if len(self.df_kmo[self.df_kmo.WC == equip_op]) != 0:
                    if kmo_on:
                        # в русской расскладке десятичный разделитель ",", в python разделитель "."
                        # схема перевода из одного вида в другой мутная, поэтому при создании
                        # таблицы КМО коэффициенты записаны целыми числами от 0 до 100
                        # этот код переводит кмо обратно в диапазон от 0 до 1
                        kmo = self.df_kmo.KMO[self.df_kmo.WC == equip_op].values[0]/100

            # дробим текущее время с учетом КМО
            cur_idx_int = current_index[0] # list to int замарочки для python
            # берем из исходных данных 1 строку с индексом current_index
            # делим время операции Time на части равные accuracy_plan
            # в конец графика добавляем копии строки исходных данных, но в поле
            # Time подставляем значение accuracy_plan, в последней страке значение может
            # быть меньше, главное проверяем, чтобы сумма времени Time добавленных строк
            # строго равнялась величине Time исходных данных. Полученные строки записываем
            # во временную таблицу df_mul
            df_mul, last_eq_mat = self.mul_row(cur_idx_int, kmo, adjust_equip, last_eq_mat, flag_gzrc)
            # Объединяем строки графика и временной талицы.
            # в python есть много команд для объединения таблиц, но pd.concat
            # рекомендована как лучшее решение, отсюда все заморочки
            self.df_graf = pd.concat([self.df_graf, df_mul.loc[:].copy()], ignore_index = True)
            # переменные bgn и end содержат индекс первой и последней строки графика
            # куда были добавлены новые строки
            end = len(self.df_graf)
            self.df_graf.reset_index(drop= True , inplace = True)
            # df_graf[bgn:end]


            ####
            rng = len(self.df_graf)
            list_range = range(bgn, end)
            #восстанавливаем данные last_seria из last_seria_mem
            last_seria = last_seria_mem
            # перебираем строки грфика в диапазоне от bgn до end
            for s in list_range:
                # s = 0
                seria = self.df_graf.iloc[s, index_col_seria]# Серия
                op = self.df_graf.iloc[s, index_col_op]# Операция
                prod = self.df_graf.iloc[s, index_col_material]# Продукция
                # заполняем поле Date
                equip_op, free_day, last_seria, gzrc_index = self.draft_equipt(s, last_seria, free_day) #df_graf,

                # ????
                seria_index = self.df_graf[(self.df_graf.Operation == op) & (self.df_graf.SeriaShort == seria) & (self.df_graf.Material == prod)].index
                if equip_op not in self.dic_gzrc and list(self.df_graf.loc[seria_index, 'Equip'])[0] != equip_op:
                    self.df_graf.loc[seria_index, 'Equip'] = equip_op
                    # for w in seria_index:
                    #     self.df_graf.iloc[w, index_col_equip] = equip_op

                self.creat_graf_row(self.df_graf, s, free_day)
                last_eq_mat[equip_op] = prod
            if equip_op == '3001 Кооперация':
                df_filter = self.df_graf.Time[self.df_graf.Equip == '3001 Кооперация']
                self.df_graf.loc[df_filter.index, 'Time'] = 0.0

            bar_length = 20 # Длина индикатора выполнения
            self.df.loc[current_index, 'Date'] = date_x
            date_x += 1
            # if date_x == 132:
            #     print('Stop')
            percent = date_x/len_df # Вычисление процента выполнения
            # Создание индикатора выполнения
            bar = "[" + "≡" * int(percent * bar_length+1) + " " * (bar_length - int(percent * bar_length+1)) + "]"
            print(f"\r{bar} {int(percent * 100)}%", f"{date_x} из {len_df}", end="")

        print("\033[0m\nГотово! Количество строк: ",  self.df_graf.shape[0])
        self.df_graf.Tday = self.pay_roll
        self.df_graf['IsKMO'] = self.df_graf.KMO != 1
        print('Загружаем из интернета производственный график...')
        self.df_calendar = self.calendar_up()
        if len(self.df_calendar) == 0:
            print('Производственный онлайн календарь недоступен\nИспользуется резервная копия')
            self.df_calendar = pd.read_excel('./WorksCalendar.xlsx')
            date_now = datetime.today()
            date_now = date_now.combine(date_now.date(), date_now.min.time())
            self.df_calendar = self.df_calendar[self.df_calendar.DateTime >= date_now]
        else:
            self.df_calendar.to_excel('./WorksCalendar.xlsx')
        # Оставляем только рабочие и короткие дни
        self.df_calendar = self.df_calendar[self.df_calendar.Type.isin(['рабочий','короткий'])]
        self.df_calendar.reset_index(inplace=True, drop=True)
        print('Записываем в график календарные дни...')
        self.df_graf['MonthDay'] = self.df_graf.Date.apply(self.graf_join_calendar)
        print('Выгружаем график в электронную таблицу...')
        self.path = self.view.envir['my_graf']
        with pd.ExcelWriter(self.path, mode="a", 
                            engine="openpyxl", 
                            if_sheet_exists='replace',
                            date_format='DD-MM-YYY') as writer:
            self.df_graf.to_excel(writer, sheet_name='График')
        sort_graf = self.df_graf[['Material', 'Cause', 'Time', 'MonthDay', 'SeriaShort', 'ExecutionTime']].groupby(['Material', 'SeriaShort', 'Cause']).agg({'ExecutionTime':'max', 'MonthDay':'max', 'Time': 'sum'})
        sort_graf.sort_values([ 'MonthDay', 'Time', 'Material'], inplace=True)
        # df_hold_parent_child = pd.DataFrame(hold_parent_child)
        # with pd.ExcelWriter(self.path, mode="a", 
        #                     engine="openpyxl", 
        #                     if_sheet_exists='replace',
        #                     date_format='DD-MM-YYY') as writer:
        #     df_hold_parent_child.to_excel(writer, sheet_name='Связи нарядов')
        return   

class Gantt:
    def __init__(self, view):
        super().__init__()
        self.df_graf = None
        self.view = view

    def load_df_graf(self):
        self.path = self.view.envir['my_graf']
        return pd.read_excel(self.path, sheet_name='График')

    def plt_gantt (self, item_gantt):#gantt_x
        #
        # Чертим диаграмму Гантта
        # gantt_x = 20
        def increment_value(value, sizes_of_groups):
            #
            # Увеличение значения счетчика value на 1 единицу
            #
            for i in range(len(value)):
                if (value[i] + 1) % sizes_of_groups[i] != 0 and sizes_of_groups[i] !=1:
                    value[i] += 1
                    return
                value[i] = 0
            pass

        def color_mix(value, sizes_of_groups):
            #
            # Создаем спектр цвета в формате r, g, b
            #
            increment_value(value, sizes_of_groups)
            r = round(value[0]*0.1, 1)
            g = round(value[1]*0.1, 1)
            b = round(value[2]*0.1, 1)
            return [r, g, b, 0.8]
        
        def creat_barh(item_gt, seria, df_graf):
            #
            #Создаем список для лини диаграммы Ганта
            #
            #Создаем bp списока из вида [0,1,2,3,4,5,6] кортеж ([0, 7])
            # item_gt = 'Equip'
            # seria = 'Верстак слесарный ЦМО №1'
            #
            barh_list = []#[первый день непрерывной серии, длина серии]

            temp_list = []#[все дни непрерывной серии]

            zero_len = len(df_graf[df_graf[item_gt] == seria].groupby(["Date"]))
            if zero_len == 0:
                return barh_list #, seria_list
            table = df_graf[['Date', 'Time']][df_graf[item_gt] == seria].groupby(by="Date").sum()#['Time']
            days = table.index
            for i in range(len(days)):
                if i == 0:
                    temp_list.append(days[i])

                elif i == len(days)-1:
                    temp_list.append(days[i])
                    barh_list.append((min(temp_list), len(temp_list)))

                elif days[i] - days[i-1] > 1:
                    barh_list.append((min(temp_list), len(temp_list)))
                    temp_list = []
                    temp_list.append(days[i])

                else:
                    temp_list.append(days[i])

            return barh_list#, seria_list

        def add_line_breaks(my_list):
            new_list = []
            for item in my_list:
                item = str(item)
                if item in self.view.envir['kolvo_smen'].keys():
                    item += f". Смен={str(self.view.envir['kolvo_smen'][item])}"
                if len(item) > 30:
                    item = item[:30] + '\n' + item[30:]
                new_list.append(item)
            return new_list

        print('Создаем диаграмму Ганта...')
        if self.df_graf is None:
            self.df_graf = self.load_df_graf()

        gantt_x = self.df_graf.Date.max()
        gui = 'Qt5Agg'
        matplotlib.use(gui, force=True)#warn=False,

        seria_list = list(self.df_graf[item_gantt].unique())
        seria_list_br = add_line_breaks(seria_list)
        # Declaring a figure "gnt"
        fig, gnt = plt.subplots(figsize=(20, 25))

        # Setting Y-axis limits
        gnt.set_ylim(0, 2*len(seria_list))

        # Setting X-axis limits
        gnt.set_xlim(1, gantt_x)#0

        # Setting labels for x-axis and y-axis
        gnt.set_xlabel('Рабочие дни')
        gnt.set_ylabel(item_gantt)

        # Setting ticks on y-axis
        gnt.set_yticks([x for x in range(2*len(seria_list)) if x%2 ==0])#[50, 100, 150, 200]
        # Labelling tickes of y-axis
        gnt.set_yticklabels(seria_list_br, fontsize=7, )#[25, 50, 75, 100] reversed()

        # Setting graph attribute
        gnt.grid(True, which='major', color='red', axis = 'x', linewidth=2)
        gnt.minorticks_on()

        gnt.grid(True, which='minor', color='grey', axis = 'x', linestyle='-', linewidth=0.5)
        gnt.grid(True, axis='x')

        # Добавляем линии диаграммы
        value = [0, 0, 0]#r, g, b инициация
        sizes_of_groups = [9, 9, 9]#r, g, b максимальные значения
        # dict_
        for k in range(len(seria_list)):
            # Declaring a bar in schedule
            c_m = color_mix(value, sizes_of_groups)
            barh_item = creat_barh(item_gantt, seria_list[k], self.df_graf)
            gnt.broken_barh(barh_item, (2*k-1, 1), facecolors = c_m)#2*len(seria_list) -

        plt.show()
        # plt.savefig("gantt1.png")
        print('Диаграмма создана')
        return  

def main():
    app = QApplication(sys.argv)
    view = View()
    model = Model(view)
    gantt = Gantt(view)
    presenter = Presenter(model, view, gantt)
    view.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()

